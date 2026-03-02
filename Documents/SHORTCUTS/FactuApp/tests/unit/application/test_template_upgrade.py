from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import json

from openpyxl import load_workbook

from app.domain.entities.invoice import Invoice
from app.domain.entities.client import Client
from app.domain.value_objects.invoice_status import InvoiceStatus
from app.infrastructure.exporters.excel_exporter_openpyxl import ExcelExporter
from app.infrastructure.exporters.pro_template_generator import generate_pro_template


def _invoice_official() -> Invoice:
    invoice = Invoice(
        id="inv-template-1",
        status=InvoiceStatus.DRAFT,
        issue_date=date(2026, 2, 18),
        client_id="client-1",
        lines=(
            Invoice.create_line("Instalacion", "1", "100.00"),
            Invoice.create_line("Mantenimiento", "2", "50.00"),
        ),
    )
    invoice.mark_official(2026, 1, datetime(2026, 2, 18, 12, 0, 0))
    return invoice


def test_pro_template_generation_creates_xlsx_and_map(tmp_path):
    template_path, map_path = generate_pro_template(tmp_path)
    assert template_path.exists()
    assert map_path.exists()

    mapping = json.loads(map_path.read_text(encoding="utf-8"))
    assert mapping["sheet"] == "Factura"
    assert "invoice_code" in mapping["cells"]


def test_template_export_with_pro_works(tmp_path):
    invoice = _invoice_official()
    generate_pro_template(tmp_path)

    exporter = ExcelExporter(export_mode="template", export_template="pro", project_root=tmp_path)
    content = exporter.render(invoice)

    wb = load_workbook(BytesIO(content))
    mapping = json.loads((tmp_path / "config" / "template_map_pro.json").read_text(encoding="utf-8"))
    ws = wb[mapping["sheet"]]
    assert ws[mapping["cells"]["invoice_code"]].value == "2026-000001"


def test_legacy_template_selection_works(tmp_path):
    invoice = _invoice_official()
    template_path, _ = generate_pro_template(tmp_path)
    legacy_path = template_path.parent / "legacy_demo.xlsx"
    legacy_path.write_bytes(template_path.read_bytes())

    exporter = ExcelExporter(export_mode="template", export_template="legacy:legacy_demo.xlsx", project_root=tmp_path)
    content = exporter.render(invoice)
    assert len(content) > 100


def test_template_failure_falls_back_to_simple(tmp_path):
    invoice = _invoice_official()
    exporter = ExcelExporter(export_mode="template", export_template="legacy:missing.xlsx", project_root=tmp_path)

    content = exporter.render(invoice)
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    assert ws["A1"].value == "Factura"


def test_template_export_uses_client_resolver_data(tmp_path):
    invoice = _invoice_official()
    generate_pro_template(tmp_path)

    client = Client(
        id="client-1",
        name="CLIENTE REAL SL",
        tax_id="B12345678",
        address="Calle Prueba 1",
        city="Madrid",
        postal_code="28000",
    )
    exporter = ExcelExporter(
        export_mode="template",
        export_template="pro",
        project_root=tmp_path,
        client_resolver=lambda client_id: client if client_id == "client-1" else None,
    )
    content = exporter.render(invoice)

    wb = load_workbook(BytesIO(content))
    mapping = json.loads((tmp_path / "config" / "template_map_pro.json").read_text(encoding="utf-8"))
    cells = mapping["cells"]
    ws = wb[mapping["sheet"]]
    assert ws[cells["client_name"]].value == "CLIENTE REAL SL"
    assert ws[cells["client_tax_id"]].value == "B12345678"
