from __future__ import annotations

import json
from collections.abc import Callable
from math import ceil
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

from app.domain.entities.invoice import Invoice
from config.settings import settings


class ExcelExporter:
    def __init__(
        self,
        export_mode: str | None = None,
        export_template: str | None = None,
        project_root: Path | None = None,
        client_resolver: Callable[[str], object | None] | None = None,
    ):
        self.export_mode = (export_mode or settings.export_mode).lower()
        self.export_template = export_template or settings.export_template
        self.project_root = project_root or Path(__file__).resolve().parents[3]
        self.client_resolver = client_resolver

    def render(self, invoice: Invoice) -> bytes:
        if self.export_mode == "simple":
            return self._render_simple(invoice)

        try:
            return self._render_template(invoice)
        except Exception:
            # Required fallback behavior: if template mode fails, fallback to simple export.
            return self._render_simple(invoice)

    def _resolve_template_and_map(self) -> tuple[Path, dict]:
        templates_excel = self.project_root / "templates" / "excel"
        config_dir = self.project_root / "config"

        if self.export_template == "pro":
            template_path = templates_excel / "factuapp_pro.xlsx"
            map_path = config_dir / "template_map_pro.json"
        elif self.export_template.startswith("legacy:"):
            file_name = self.export_template.split(":", 1)[1]
            template_path = templates_excel / file_name
            legacy_map = templates_excel / f"{Path(file_name).stem}.map.json"
            if legacy_map.exists():
                map_path = legacy_map
            else:
                map_path = config_dir / "template_map_legacy_invoice.json"
        else:
            template_path = templates_excel / "factuapp_pro.xlsx"
            map_path = config_dir / "template_map_pro.json"

        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        if not map_path.exists():
            raise FileNotFoundError(f"Template map not found: {map_path}")

        mapping = json.loads(map_path.read_text(encoding="utf-8"))
        return template_path, mapping

    def _resolve_logo_path(self) -> Path | None:
        templates_logo_dir = self.project_root / "templates" / "logo"
        assets_dir = self.project_root / "assets"

        named_candidates = [
            templates_logo_dir / "LOGO_HORIZONTAL.png",
            templates_logo_dir / "logo.png",
            assets_dir / "logo.png",
            templates_logo_dir / "logo.jpg",
            templates_logo_dir / "logo.jpeg",
            assets_dir / "logo.jpg",
            assets_dir / "logo.jpeg",
        ]
        path = next((candidate for candidate in named_candidates if candidate.exists()), None)
        if path:
            return path

        for ext in ("*.png", "*.jpg", "*.jpeg"):
            generic = sorted(templates_logo_dir.glob(ext))
            if generic:
                return generic[0]
        return None

    def _inject_logo(self, ws, anchor: str) -> None:
        if getattr(ws, "_images", None):
            return
        logo = self._resolve_logo_path()
        if not logo:
            return
        try:
            img = XLImage(str(logo))
            if anchor == "A1":
                # Expand row 1 so the logo is clearly visible as a proper header area.
                # Logo ratio 468×149 ≈ 3.14:1; at 54pt row height → ~72px → width ≈ 226px.
                ws.row_dimensions[1].height = 54
                img.width = 210
                img.height = 67
            else:
                img.width = 160
                img.height = 51
            ws.add_image(img, anchor)
        except Exception:
            pass

    def _resolve_client_fields(self, invoice: Invoice) -> dict[str, str]:
        default_name = invoice.client_id or "Cliente no definido"
        if not invoice.client_id or not self.client_resolver:
            return {
                "name": default_name,
                "tax_id": "",
                "address": "",
                "city": "",
                "postal_code": "",
            }

        client = self.client_resolver(invoice.client_id)
        if client is None:
            return {
                "name": default_name,
                "tax_id": "",
                "address": "",
                "city": "",
                "postal_code": "",
            }

        return {
            "name": getattr(client, "name", None) or default_name,
            "tax_id": getattr(client, "tax_id", None) or "",
            "address": getattr(client, "address", None) or "",
            "city": getattr(client, "city", None) or "",
            "postal_code": getattr(client, "postal_code", None) or "",
        }

    def _format_invoice_code(self, invoice: Invoice, mapping: dict) -> str:
        code = invoice.official_code or invoice.id
        if mapping.get("sheet") == "Invoice Template":
            return f"Factura Nº {code}"
        return code

    def _fill_issuer_fields(self, ws, cells: dict) -> None:
        issuer_values = {
            "issuer_name": settings.issuer_name,
            "issuer_tax_id": f"CIF: {settings.issuer_tax_id}",
            "issuer_address": settings.issuer_address,
            "issuer_email": settings.issuer_email,
            "issuer_phone": settings.issuer_phone,
            "issuer_country_currency": f"{settings.issuer_country} · {settings.currency}",
            "iban": f"IBAN: {settings.issuer_iban}",
        }
        for key, value in issuer_values.items():
            if key in cells:
                ws[cells[key]] = value

    @staticmethod
    def _estimate_row_height(description: str, base_height: float = 18) -> float:
        if not description:
            return base_height
        wrapped_lines = max(1, ceil(len(description) / 58))
        return max(base_height, wrapped_lines * 16)

    def _render_template(self, invoice: Invoice) -> bytes:
        template_path, mapping = self._resolve_template_and_map()
        wb = load_workbook(template_path)
        ws = wb[mapping["sheet"]]
        cells = mapping["cells"]
        table = mapping["table"]

        logo_anchor = cells.get("logo_anchor")
        if logo_anchor:
            self._inject_logo(ws, logo_anchor)
        self._fill_issuer_fields(ws, cells)

        client = self._resolve_client_fields(invoice)
        ws[cells["invoice_code"]] = self._format_invoice_code(invoice, mapping)
        date_cell = ws[cells["issue_date"]]
        date_cell.value = invoice.issue_date
        date_cell.number_format = "DD/MM/YYYY"
        ws[cells["client_name"]] = client["name"]
        ws[cells["client_tax_id"]] = client["tax_id"]
        ws[cells["client_address"]] = client["address"]
        if "client_city" in cells:
            ws[cells["client_city"]] = client["city"]
        if "client_postal" in cells:
            ws[cells["client_postal"]] = client["postal_code"]

        ws[cells["notes"]] = ""
        ws[cells["hash"]] = ""
        if "subtotal_label" in cells:
            ws[cells["subtotal_label"]] = "Base"
        if "tax_label" in cells:
            ws[cells["tax_label"]] = f"IVA {int(settings.vat_rate * 100)}%"
        if "total_label" in cells:
            ws[cells["total_label"]] = "TOTAL"

        ws[cells["subtotal_value"]] = float(invoice.subtotal)
        ws[cells["tax_value"]] = float(invoice.tax_amount)
        ws[cells["total_value"]] = float(invoice.total)

        start_row = int(table["start_row"])
        max_rows = int(table["max_rows"])
        columns = table["columns"]

        for i in range(max_rows):
            row = start_row + i
            ws[f"{columns['description']}{row}"] = None
            ws[f"{columns['qty']}{row}"] = None
            ws[f"{columns['unit_price']}{row}"] = None
            ws[f"{columns['subtotal']}{row}"] = None
            if "unit" in columns:
                ws[f"{columns['unit']}{row}"] = None

        for idx, line in enumerate(invoice.lines[:max_rows]):
            row = start_row + idx
            ws[f"{columns['description']}{row}"] = line.description
            ws[f"{columns['qty']}{row}"] = float(line.quantity)
            ws[f"{columns['unit_price']}{row}"] = float(line.unit_price)
            ws[f"{columns['subtotal']}{row}"] = float(line.line_subtotal)
            if "unit" in columns:
                ws[f"{columns['unit']}{row}"] = "ud"
            if mapping.get("sheet") == "Factura":
                ws.row_dimensions[row].height = self._estimate_row_height(line.description)

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    def _render_simple(self, invoice: Invoice) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Factura"
        self._inject_logo(ws, "A1")
        ws.append(["Factura", invoice.official_code or invoice.id])
        ws.append(["Estado", invoice.status.value])
        ws.append(["Fecha", invoice.issue_date.isoformat()])
        ws.append([])
        ws.append(["Descripcion", "Cantidad", "Precio", "Subtotal", "IVA", "Total"])
        for line in invoice.lines:
            ws.append([
                line.description,
                str(line.quantity),
                str(line.unit_price),
                str(line.line_subtotal),
                str(line.line_tax),
                str(line.line_total),
            ])
        ws.append([])
        ws.append(["Subtotal", str(invoice.subtotal)])
        ws.append(["IVA", str(invoice.tax_amount)])
        ws.append(["Total", str(invoice.total)])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
