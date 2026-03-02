from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


DEFAULT_MAP = {
    "sheet": "Factura",
    "cells": {
        "logo_anchor": "B2",
        "issuer_name": "F2",
        "issuer_tax_id": "F3",
        "issuer_address": "F4",
        "issuer_email": "F5",
        "issuer_phone": "F6",
        "issuer_country_currency": "F7",
        "client_title": "B9",
        "client_name": "B10",
        "client_tax_id": "B11",
        "client_address": "B12",
        "invoice_title": "B7",
        "invoice_code": "D7",
        "issue_date": "I7",
        "table_header_description": "B15",
        "table_header_qty": "F15",
        "table_header_unit": "G15",
        "table_header_unit_price": "H15",
        "table_header_subtotal": "I15",
        "subtotal_label": "G37",
        "subtotal_value": "I37",
        "tax_label": "G38",
        "tax_value": "I38",
        "total_label": "G39",
        "total_value": "I39",
        "iban": "B37",
        "notes": "B38",
        "hash": "B39",
    },
    "table": {
        "start_row": 16,
        "max_rows": 20,
        "columns": {
            "description": "B",
            "qty": "F",
            "unit": "G",
            "unit_price": "H",
            "subtotal": "I",
        },
    },
}


def _write_logo_or_placeholder(ws, assets_dir: Path, templates_logo_dir: Path) -> None:
    png_candidates = [
        templates_logo_dir / "LOGO_HORIZONTAL.png",
        templates_logo_dir / "logo.png",
        assets_dir / "logo.png",
    ]
    svg_candidates = [
        templates_logo_dir / "logo.svg",
        assets_dir / "logo.svg",
    ]

    png_logo = next((candidate for candidate in png_candidates if candidate.exists()), None)
    svg_logo = next((candidate for candidate in svg_candidates if candidate.exists()), None)

    if png_logo:
        image = XLImage(str(png_logo))
        image.width = 200
        image.height = 62
        ws.add_image(image, "B2")
        return

    ws["B2"] = "[LOGO]"
    ws["B2"].font = Font(size=14, bold=True, name="Calibri")
    ws["B2"].alignment = Alignment(horizontal="left")
    if svg_logo is not None:
        ws["B3"] = "Logo SVG detectado (openpyxl no incrusta SVG). Exporta PNG a assets/logo.png"
        ws["B3"].font = Font(size=9, italic=True, name="Calibri")


def generate_pro_template(project_root: Path) -> tuple[Path, Path]:
    templates_excel = project_root / "templates" / "excel"
    config_dir = project_root / "config"
    assets_dir = project_root / "assets"
    templates_excel.mkdir(parents=True, exist_ok=True)
    config_dir.mkdir(parents=True, exist_ok=True)

    template_path = templates_excel / "factuapp_pro.xlsx"
    map_path = config_dir / "template_map_pro.json"

    wb = Workbook()
    ws = wb.active
    ws.title = DEFAULT_MAP["sheet"]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16

    for row in range(1, 50):
        ws.row_dimensions[row].height = 18

    templates_logo_dir = project_root / "templates" / "logo"
    _write_logo_or_placeholder(ws, assets_dir, templates_logo_dir)

    ws["F2"] = "Arteclima 2020 SLU"
    ws["F3"] = "CIF: B88644455"
    ws["F4"] = "Calle Río Guadarrama 22, 28971, Griñón, Madrid"
    ws["F5"] = "arteclima@arteclimaslu.com"
    ws["F6"] = "+34 637 568 249"
    ws["F7"] = "Spain · EUR"
    ws["B9"] = "CLIENTE"
    ws["B10"] = "Nombre cliente"
    ws["B11"] = "CIF/NIF"
    ws["B12"] = "Dirección cliente"

    ws["B7"] = "FACTURA Nº"
    ws["D7"] = "2026-000001"
    ws["I7"] = "18/02/2026"

    header_fill = PatternFill(fill_type="solid", start_color="E6E6E6", end_color="E6E6E6")
    header_font = Font(color="000000", bold=True, name="Calibri")
    border = Border(
        left=Side(style="thin", color="C7D0D9"),
        right=Side(style="thin", color="C7D0D9"),
        top=Side(style="thin", color="C7D0D9"),
        bottom=Side(style="thin", color="C7D0D9"),
    )

    ws["B15"] = "Descripción"
    ws["F15"] = "Cantidad"
    ws["G15"] = "Unidad"
    ws["H15"] = "Precio Unitario"
    ws["I15"] = "Subtotal"

    for cell in ["B15", "F15", "G15", "H15", "I15"]:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].border = border

    table_start = DEFAULT_MAP["table"]["start_row"]
    max_rows = DEFAULT_MAP["table"]["max_rows"]
    for row in range(table_start, table_start + max_rows):
        for col in ["B", "F", "G", "H", "I"]:
            c = ws[f"{col}{row}"]
            c.border = border
            c.font = Font(name="Calibri", size=11)
            if col == "B":
                c.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                c.alignment = Alignment(horizontal="right", vertical="top")
            if col in {"H", "I"}:
                c.number_format = '#,##0.00 [$€-x-euro2]'
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "B16"

    ws["G37"] = "Base"
    ws["G38"] = "IVA 21%"
    ws["G39"] = "TOTAL"
    ws["I37"] = 0
    ws["I38"] = 0
    ws["I39"] = 0

    for cell in ["I37", "I38", "I39"]:
        ws[cell].number_format = '#,##0.00 [$€-x-euro2]'
        ws[cell].alignment = Alignment(horizontal="right")

    ws["G39"].font = Font(bold=True, name="Calibri", size=12)
    ws["I39"].font = Font(bold=True, size=14, name="Calibri")

    ws["B37"] = "IBAN: ES78 0081 5149 8100 0128 2333"
    ws["B38"] = "Notas:"
    ws["B39"] = "Hash:"
    ws["B39"].font = Font(size=9, color="5A6B7B", name="Calibri")

    # Header, client and totals visual blocks tuned for A4->PDF conversion.
    for row in range(2, 8):
        for col in ("F", "G", "H", "I"):
            c = ws[f"{col}{row}"]
            c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            c.font = Font(name="Calibri", size=10, bold=(row == 2))
    for cell in ("B7", "D7", "I7"):
        ws[cell].font = Font(name="Calibri", bold=True, size=14 if cell == "D7" else 12)
    for row in range(9, 13):
        for col in ("B", "C", "D", "E", "F", "G", "H", "I"):
            c = ws[f"{col}{row}"]
            c.border = border
            c.font = Font(name="Calibri", size=10, bold=(row == 9))
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    wb.save(template_path)
    map_path.write_text(json.dumps(DEFAULT_MAP, indent=2, ensure_ascii=False), encoding="utf-8")
    return template_path, map_path


def ensure_pro_template(project_root: Path) -> None:
    template_path = project_root / "templates" / "excel" / "factuapp_pro.xlsx"
    map_path = project_root / "config" / "template_map_pro.json"
    if template_path.exists() and map_path.exists():
        return
    generate_pro_template(project_root)
